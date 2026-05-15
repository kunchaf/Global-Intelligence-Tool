import csv
from io import BytesIO, StringIO
from typing import Tuple

from .compare import DIM_LABELS


def build_compare_excel(payload: dict) -> Tuple[BytesIO, str, str]:
    """
    Build a spreadsheet for download.
    Returns (buffer, download_filename, mimetype).
    Uses .xlsx when openpyxl is installed; otherwise UTF-8 CSV (opens in Excel).
    """
    try:
        return _build_xlsx(payload)
    except ImportError:
        return _build_csv(payload)


def _build_csv(payload: dict) -> Tuple[BytesIO, str, str]:
    dimensions = payload.get("dimensions") or []
    rows = payload.get("rows") or []

    headers = [
        "Country",
        "Your input",
        "ISO code",
        "Latitude",
        "Longitude",
    ] + [DIM_LABELS.get(d, d) for d in dimensions]

    buf = StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)

    for row in rows:
        if row.get("error"):
            out = ["", row.get("input", ""), "", "", ""]
            out.extend(
                [row.get("error", "") if i == 0 else "" for i in range(len(dimensions))]
            )
            writer.writerow(out)
            continue
        line = [
            row.get("country") or "",
            row.get("input") or "",
            row.get("country_code") or "",
            row.get("latitude") if row.get("latitude") is not None else "",
            row.get("longitude") if row.get("longitude") is not None else "",
        ]
        for d in dimensions:
            v = row.get(d)
            line.append("" if v is None else str(v))
        writer.writerow(line)

    bio = BytesIO(buf.getvalue().encode("utf-8-sig"))
    bio.seek(0)
    return (
        bio,
        "demography_compare.csv",
        "text/csv; charset=utf-8",
    )


def _build_xlsx(payload: dict) -> Tuple[BytesIO, str, str]:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    dimensions = payload.get("dimensions") or []
    rows = payload.get("rows") or []

    wb = Workbook()
    ws = wb.active
    ws.title = "Compare"

    headers = [
        "Country",
        "Your input",
        "ISO code",
        "Latitude",
        "Longitude",
    ] + [DIM_LABELS.get(d, d) for d in dimensions]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        if row.get("error"):
            out = ["", row.get("input", ""), "", "", ""]
            out.extend(
                [row.get("error", "") if i == 0 else "" for i in range(len(dimensions))]
            )
            ws.append(out)
            continue

        line = [
            row.get("country") or "",
            row.get("input") or "",
            row.get("country_code") or "",
            row.get("latitude") if row.get("latitude") is not None else "",
            row.get("longitude") if row.get("longitude") is not None else "",
        ]
        for d in dimensions:
            v = row.get(d)
            line.append("" if v is None else str(v))
        ws.append(line)

    for cells in ws.columns:
        letter = get_column_letter(cells[0].column)
        max_len = 0
        for cell in cells:
            try:
                max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
            except Exception:
                pass
        ws.column_dimensions[letter].width = min(max_len + 2, 60)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return (
        bio,
        "demography_compare.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
